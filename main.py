import openpyxl
import tkinter.filedialog
import pandas as pd

file_template_path = tkinter.filedialog.askopenfilename(title='MỞ FILE TEMPLATE')
file_vlookup = tkinter.filedialog.askopenfilename(title='MỞ FILE VLOOKUP')

df_vlookup = pd.read_excel(file_vlookup)
df_vlookup['Mã trang thiết bị new'] = df_vlookup['Mã trang thiết bị new'].str.replace('\n', '; ')
df_vlookup['Gộp'] = df_vlookup['Tên trang thiết bị'] + ' (' + df_vlookup['Mã trang thiết bị new'] + ')'

template_print_wb =openpyxl.load_workbook(file_template_path)
originalsheet =template_print_wb['Original Sheet']

for i in set(df_vlookup['Gộp']):
    duplicate_sheet =template_print_wb.copy_worksheet(originalsheet)
    duplicate_sheet.title = i[:15]
    template_print_wb.active = duplicate_sheet
    row_index = 0
    for row in duplicate_sheet.iter_rows(values_only=True):
        break_point = False
        for col_index in range(len(row)):
            if 'Mã thiết bị' in str(row[col_index]):
                duplicate_sheet.cell(row=row_index + 1, column=col_index + 1, value="Mã thiết bị: " + str(i))
                break_point = True
                break
        if break_point:
            break
        row_index = row_index + 1

template_print_wb.save('fileprint.xlsx')

file_print = openpyxl.load_workbook('fileprint.xlsx')
sheet_count = len(file_print.sheetnames)
print(sheet_count)


